"""统计页「已验证看板」导出 Excel（按 verified_markdown 中表格合并单元格写入）。"""

from __future__ import annotations

import re
from io import BytesIO
from typing import Any

from bs4 import BeautifulSoup, Tag
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

_align_wrap = Alignment(vertical="center", wrap_text=True)
_align_header = Alignment(vertical="center", horizontal="center", wrap_text=True)

# 金加 CNC0 导出：该行与摘要首行相同，每格占两列（Excel 行号 1-based）
CNC0_PAIR_COLS_EXCEL_ROW = 23


def _table_cell_compact(value: str) -> str:
    """与入库解析侧 _compact_text 规则一致，压缩单元格内空白。"""
    return re.sub(r"\s+", " ", value or "").strip()


def _write_verified_html_table_to_sheet(
    ws: Any,
    table: Tag,
    start_row: int,
    *,
    force_first_row_pair_cols: bool = False,
    extra_pair_cols_excel_rows: frozenset[int] | None = None,
) -> int:
    """把 verified_markdown 里单张 <table> 写入 Sheet。

    force_first_row_pair_cols：摘要表首行每格占两列。
    extra_pair_cols_excel_rows：额外指定若干 **Excel 绝对行号**，该行每格也占两列（如 CNC0 第 23 行）。
    其余行按 HTML rowspan/colspan。
    """
    rows = table.find_all("tr")
    if not rows:
        return start_row

    extra_pairs = extra_pair_cols_excel_rows or frozenset()

    occ: list[list[bool]] = []

    def _ensure(r: int, c: int) -> None:
        while len(occ) <= r:
            occ.append([])
        while len(occ[r]) <= c:
            occ[r].append(False)

    def _is_blocked(r: int, c: int) -> bool:
        _ensure(r, c)
        return occ[r][c]

    def _mark(r0: int, rs: int, c0: int, cs: int) -> None:
        for dr in range(rs):
            for dc in range(cs):
                _ensure(r0 + dr, c0 + dc)
                occ[r0 + dr][c0 + dc] = True

    out_row = start_row
    for row_idx, row in enumerate(rows):
        cells = row.find_all(["td", "th"])
        col_idx = 0
        for cell in cells:
            while _is_blocked(row_idx, col_idx):
                col_idx += 1
            rs = max(1, int(cell.get("rowspan") or 1))
            rs = min(rs, len(rows) - row_idx)
            r1 = out_row + row_idx
            force_pair = (row_idx == 0 and force_first_row_pair_cols) or (r1 in extra_pairs)
            if force_pair:
                cs = 2
            else:
                cs = max(1, int(cell.get("colspan") or 1))
            value = _table_cell_compact(cell.get_text(" ", strip=True))
            c1 = col_idx + 1
            scl = ws.cell(row=r1, column=c1, value=value or None)
            scl.alignment = _align_header if cell.name == "th" else _align_wrap
            if rs > 1 or cs > 1:
                ws.merge_cells(
                    start_row=r1,
                    end_row=r1 + rs - 1,
                    start_column=c1,
                    end_column=c1 + cs - 1,
                )
            _mark(row_idx, rs, col_idx, cs)
            col_idx += cs

    return out_row + len(rows)


def _column_width_from_text(text: str) -> float:
    """按中英混排粗算列宽（Excel 单位），用于数据行字数自适应。"""
    units = 0.0
    for c in str(text):
        if "\u4e00" <= c <= "\u9fff" or c in "，。；：、（）「」『』":
            units += 1.85
        elif c == " ":
            units += 0.5
        else:
            units += 1.0
    return max(7.0, min(48.0, units + 1.4))


def _fit_verified_sheet_columns(ws: Any, *, skip_rows: set[int] | None = None) -> None:
    """列宽只按「非表头行」单元格内容估算（字数自适应）；不写表头行避免把列撑得过宽。"""
    skip_rows = skip_rows or set()
    if not ws.max_column:
        return
    for col_idx in range(1, ws.max_column + 1):
        width_need = 8.0
        for row_idx in range(1, ws.max_row + 1):
            if row_idx in skip_rows:
                continue
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value is None:
                continue
            width_need = max(width_need, _column_width_from_text(str(cell.value)))
        ws.column_dimensions[get_column_letter(col_idx)].width = round(width_need, 2)


def _unique_excel_sheet_name(base: str, used: set[str]) -> str:
    raw = "".join("_" if c in "[]:*?/\\" else c for c in str(base))
    raw = (raw.strip() or "task")[:31]
    name = raw
    i = 2
    while name in used:
        suffix = f"_{i}"
        name = (raw[: max(1, 31 - len(suffix))] + suffix)[:31]
        i += 1
    used.add(name)
    return name


def _production_date_iso(rec: dict[str, Any]) -> str:
    d = rec["production_date"]
    if hasattr(d, "isoformat"):
        return d.isoformat()
    return str(d)


def build_verified_board_excel_bytes(
    board_rows: list[dict[str, Any]],
    markdown_by_ocr_id: dict[int, str | None],
) -> bytes:
    """按看板行生成多 sheet（按 task_id 升序，标签页从左到右），表名为 task_id - 生产日期。"""
    def _part_sort(r: dict[str, Any]) -> int:
        p = r.get("part")
        return int(p) if p is not None else 0

    ordered = sorted(
        board_rows,
        key=lambda r: (
            r["task_id"],
            r["production_date"],
            _part_sort(r),
            r["ocr_result_id"],
        ),
    )
    wb = Workbook()
    wb.remove(wb.active)
    used_titles: set[str] = set()
    for rec in ordered:
        task_id = rec["task_id"]
        ocr_id = rec["ocr_result_id"]
        dstr = _production_date_iso(rec)
        base_title = f"{task_id} - {dstr}"
        ws = wb.create_sheet(title=_unique_excel_sheet_name(base_title, used_titles))
        html = (markdown_by_ocr_id.get(ocr_id) or "").strip()
        if not html:
            ws.cell(row=1, column=1, value="（无 verified_markdown）")
            _fit_verified_sheet_columns(ws, skip_rows=set())
            continue
        soup = BeautifulSoup(f"<div>{html}</div>", "html.parser")
        tables = soup.find_all("table")
        if not tables:
            ws.cell(row=1, column=1, value="（无表格内容）")
            _fit_verified_sheet_columns(ws, skip_rows=set())
            continue
        row_cursor = 1
        table_header_rows: set[int] = set()
        is_cnc0 = rec.get("process_name") == "CNC0"
        cnc0_extra_pairs = frozenset({CNC0_PAIR_COLS_EXCEL_ROW}) if is_cnc0 else frozenset()
        if is_cnc0:
            table_header_rows.add(CNC0_PAIR_COLS_EXCEL_ROW)
        for ti, table in enumerate(tables):
            table_header_rows.add(row_cursor)
            row_cursor = _write_verified_html_table_to_sheet(
                ws,
                table,
                row_cursor,
                force_first_row_pair_cols=(ti == 0),
                extra_pair_cols_excel_rows=cnc0_extra_pairs,
            )
            row_cursor += 1
        _fit_verified_sheet_columns(ws, skip_rows=table_header_rows)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
