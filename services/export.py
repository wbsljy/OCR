
"""看板：按条件查表得到 DataFrame，原样写回 Excel。"""

from __future__ import annotations

from collections.abc import Sequence
from datetime import date, timedelta
from io import BytesIO
from typing import Any

import pandas as pd
from sqlalchemy import select
from sqlalchemy.orm import Session
# 添加边框
from openpyxl.styles import Border, Side
from openpyxl.styles import Alignment
from services.dashboard_service import (
    CNC0_DEFECT_TYPES,
    CNC0_FULL_DEFECT_TYPES,
    DEFAULT_PRODUCT_NAME,
    DUANYA_DEFECT_TYPES,
    GURONG_DEFECT_TYPES,
    INSPECTION_LOCATION_OPTIONS,
    KEY_PROCESS_OPTIONS,
    PROCESS_MODEL_MAP,
    PRODUCT_NAME_OPTIONS,
    SHIFT_OPTIONS,
    SHIXIAO_DEFECT_TYPES,
    parse_iso_date,
)
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.layout import Layout, ManualLayout
from openpyxl.chart.label import DataLabel, DataLabelList
from openpyxl.chart.text import RichText
from openpyxl.drawing.text import (
    CharacterProperties,
    Paragraph,
    ParagraphProperties,
    RichTextProperties,
)
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, TwoCellAnchor


def _finalize_dashboard_rate_chart(final_data: pd.DataFrame,bar: BarChart, ws: Any) -> None:
    """缩小绘图区、略增大表上图表占位，减轻标题/坐标轴/底部图例在 Excel 中重叠。"""
    bar.layout = Layout(
        manualLayout=ManualLayout(
            layoutTarget="outer",
            xMode="edge",
            yMode="edge",
            wMode="edge",
            hMode="edge",
            x=0.06,
            y=0.13,
            w=0.88,
            h=0.85,
        )
    )
    bar.anchor = TwoCellAnchor(
        _from=AnchorMarker(col=0, colOff=0, row=1, rowOff=0),
        to=AnchorMarker(col=len(final_data.columns)+1, colOff=0, row=10, rowOff=0),
    )
    ws.add_chart(bar)


def export_rows_is_empty(rows: pd.DataFrame | Sequence[Any]) -> bool:
    if isinstance(rows, pd.DataFrame):
        return rows.empty
    return len(rows) == 0


def fetch_dashboard_records_for_export(
    db: Session,
    *,
    key_name: str | None,
    process_name: str | None,
    start_date: str | None,
    end_date: str | None,
    production_name: str | None = None,
    inspection_location: str | None = None,
    shift_name: str | None = None,
) -> tuple[str, str, date, date, pd.DataFrame]:
    selected_key = key_name if key_name in KEY_PROCESS_OPTIONS else "沖壓"
    process_options = KEY_PROCESS_OPTIONS[selected_key]
    selected_process = process_name if process_name in process_options else process_options[0]
    pair = (selected_key, selected_process)
    if pair not in PROCESS_MODEL_MAP:
        raise ValueError("无效的 key / 製程组合。")

    end_value = parse_iso_date(end_date) or date.today()
    start_value = parse_iso_date(start_date) or (end_value - timedelta(days=29))
    if start_value > end_value:
        start_value, end_value = end_value, start_value

    selected_production_name = (
        production_name if production_name in PRODUCT_NAME_OPTIONS else DEFAULT_PRODUCT_NAME
    )
    selected_inspection_location = (
        inspection_location
        if inspection_location in INSPECTION_LOCATION_OPTIONS
        else "不限"
    )
    selected_shift = shift_name if shift_name in SHIFT_OPTIONS else "不限"

    model = PROCESS_MODEL_MAP[pair]
    stmt = (
        select(model)
        .where(
            model.production_date >= start_value,
            model.production_date <= end_value,
            model.product_name == selected_production_name,
        )
        .order_by(model.production_date.asc())
    )
    if selected_shift != "不限":
        stmt = stmt.where(model.shift == selected_shift)
    if selected_process == "CNC0" :
        stmt = stmt.where(model.inspection_location == selected_inspection_location)
    df = pd.read_sql(stmt, db.get_bind())
    return selected_key, selected_process, start_value, end_value, df

def _rate_to_float(value: Any) -> float:
    if isinstance(value, str):
        v = value.strip()
        if v.endswith("%"):
            try:
                return float(v[:-1]) / 100
            except ValueError:
                return 0.0
        try:
            return float(v)
        except ValueError:
            return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    return 0.0


def add_native_good_rate_chart_choya(ws, final_data: pd.DataFrame, anchor: str = "A2") -> None:
    """
    使用 openpyxl 原生图表生成组合图（汇总柱状 + 日趋势折线 + 目标线）。
    """
    def _text_props(size_pt: float) -> RichText:
        size = int(size_pt * 100)
        return RichText(
            bodyPr=RichTextProperties(),
            p=[
                Paragraph(
                    pPr=ParagraphProperties(defRPr=CharacterProperties(sz=size)),
                    endParaRPr=CharacterProperties(sz=size),
                )
            ],
        )

    date_labels = [str(col) for col in final_data.columns.tolist()[1:]]
    if not date_labels:
        return

    total_rate_values = [_rate_to_float(v) for v in final_data.iloc[4][1:].tolist()]
    if not total_rate_values:
        return

    # 查找日数据分界：例如 04/21 视为日趋势部分
    split_index = next((i for i, label in enumerate(date_labels) if "/" in label), len(date_labels))
    if split_index == 0:
        split_index = len(date_labels)

    target_values = [0.998] * len(total_rate_values)
    summary_values = [
        total_rate_values[i] if i < split_index else None for i in range(len(total_rate_values))
    ]
    daily_values = [
        total_rate_values[i] if i >= split_index else None for i in range(len(total_rate_values))
    ]

    # 在右侧隐藏区域准备图表数据源，避免影响主展示区域
    helper_col = 70  # BR 列附近
    ws.cell(row=2, column=helper_col, value="序列")
    ws.cell(row=2, column=helper_col + 1, value="日期")
    ws.cell(row=3, column=helper_col, value="年/月汇总")
    ws.cell(row=4, column=helper_col, value="日良率趋势")
    ws.cell(row=5, column=helper_col, value="目标良率")

    for idx, (label, summary, daily, target) in enumerate(
        zip(date_labels, summary_values, daily_values, target_values)
    ):
        col = helper_col + 1 + idx
        ws.cell(row=2, column=col, value=label)
        ws.cell(row=3, column=col, value=summary)
        ws.cell(row=4, column=col, value=daily)
        ws.cell(row=5, column=col, value=target)
        ws.cell(row=3, column=col).number_format = "0.00%"
        ws.cell(row=4, column=col).number_format = "0.00%"
        ws.cell(row=5, column=col).number_format = "0.00%"

    end_col = helper_col + len(date_labels)
    cats = Reference(ws, min_col=helper_col + 1, max_col=end_col, min_row=2, max_row=2)

    # 柱状图：年/月汇总
    bar = BarChart()
    bar.title = "总良率趋势图"
    bar.type = "col"
    bar.style = 10
    bar.y_axis.title = "总良率"
    bar.x_axis.title = "日期"
    # 尺寸将由 TwoCellAnchor 控制，这里仅给默认值
    bar.height = 5.0
    bar.width = 14.0
    bar.gapWidth = 140
    bar.y_axis.number_format = "0.00%"
    bar.title.txPr = _text_props(10)
    bar.x_axis.txPr = _text_props(8)
    bar.y_axis.txPr = _text_props(8)

    min_rate = min(total_rate_values + [0.998])
    bar.y_axis.scaling.min = max(0.90, min_rate - 0.01)
    bar.y_axis.scaling.max = 1.00

    bar_data = Reference(ws, min_col=helper_col, max_col=end_col, min_row=3, max_row=3)
    bar.add_data(bar_data, titles_from_data=True, from_rows=True)
    bar.set_categories(cats)
    bar.x_axis.delete = False
    bar.y_axis.delete = False
    bar.x_axis.tickLblPos = "nextTo"
    bar.y_axis.tickLblPos = "nextTo"
    if bar.series:
        bar.series[0].graphicalProperties.solidFill = "8EC7F7"
        bar.series[0].graphicalProperties.line.solidFill = "4A90E2"
        bar.series[0].dLbls = DataLabelList()
        bar.series[0].dLbls.showVal = True
        bar.series[0].dLbls.showSerName = False
        bar.series[0].dLbls.showCatName = False
        bar.series[0].dLbls.showLegendKey = False
        bar.series[0].dLbls.showPercent = False
        bar.series[0].dLbls.dLblPos = "outEnd"
        bar.series[0].dLbls.numFmt = "0.00%"
        bar.series[0].dLbls.txPr = _text_props(7)

    # 折线：日趋势 + 目标线
    line = LineChart()
    line_data = Reference(ws, min_col=helper_col, max_col=end_col, min_row=4, max_row=5)
    line.add_data(line_data, titles_from_data=True, from_rows=True)
    line.set_categories(cats)
    line.y_axis.axId = 200
    line.y_axis.scaling.min = bar.y_axis.scaling.min
    line.y_axis.scaling.max = 1.00
    line.y_axis.number_format = "0.00%"
    line.y_axis.majorGridlines = None
    line.y_axis.delete = True

    if len(line.series) >= 1:
        # 日趋势
        line.series[0].graphicalProperties.line.solidFill = "F39C12"
        line.series[0].graphicalProperties.line.width = 28575  # 2.25pt
        line.series[0].marker.symbol = "circle"
        line.series[0].marker.size = 6
        line.series[0].dLbls = DataLabelList()
        line.series[0].dLbls.showVal = True
        line.series[0].dLbls.showSerName = False
        line.series[0].dLbls.showCatName = False
        line.series[0].dLbls.showLegendKey = False
        line.series[0].dLbls.showPercent = False
        line.series[0].dLbls.dLblPos = "t"
        line.series[0].dLbls.numFmt = "0.00%"
        line.series[0].dLbls.txPr = _text_props(7)
    if len(line.series) >= 2:
        # 目标线
        line.series[1].graphicalProperties.line.solidFill = "D9534F"
        line.series[1].graphicalProperties.line.width = 19050  # 1.5pt
        line.series[1].graphicalProperties.line.dashStyle = "sysDot"
        line.series[1].marker.symbol = "triangle"
        line.series[1].marker.size = 5
        line.series[1].dLbls = DataLabelList()
        line.series[1].dLbls.showVal = False
        line.series[1].dLbls.showSerName = False
        line.series[1].dLbls.showCatName = False
        line.series[1].dLbls.showLegendKey = False
        line.series[1].dLbls.showPercent = False
        line.series[1].dLbls.dLblPos = "r"
        line.series[1].dLbls.numFmt = "0.00%"
        line.series[1].dLbls.txPr = _text_props(7)
        line.series[1].dLbls.dLbl = [
            DataLabel(
                idx=max(0, len(target_values) - 1),
                showVal=True,
                showSerName=False,
                showCatName=False,
                showLegendKey=False,
                showPercent=False,
                showBubbleSize=False,
                dLblPos="r",
                numFmt="0.00%",
                separator="",
            )
        ]

    bar += line
    bar.legend.position = "b"
    bar.legend.txPr = _text_props(7.5)
    _finalize_dashboard_rate_chart(final_data,bar, ws)

def add_native_good_rate_chart_jinjia_cnc0_full(ws, final_data: pd.DataFrame, anchor: str = "A2") -> None:
    """
    openpyxl 组合图：双柱（一次/二次年或月汇总）+ 四线（两次日趋势 + 两项目标良率）。
    """
    def _text_props(size_pt: float) -> RichText:
        size = int(size_pt * 100)
        return RichText(
            bodyPr=RichTextProperties(),
            p=[
                Paragraph(
                    pPr=ParagraphProperties(defRPr=CharacterProperties(sz=size)),
                    endParaRPr=CharacterProperties(sz=size),
                )
            ],
        )

    date_labels = [str(col) for col in final_data.columns.tolist()[1:]]
    if not date_labels:
        return

    first_target_values = [_rate_to_float(v) for v in final_data.iloc[2][1:].tolist()]
    second_target_values = [_rate_to_float(v) for v in final_data.iloc[3][1:].tolist()]
    first_rate_values = [_rate_to_float(v) for v in final_data.iloc[7][1:].tolist()]
    second_rate_values = [_rate_to_float(v) for v in final_data.iloc[8][1:].tolist()]
    n = len(date_labels)
    if (
        not first_rate_values
        or len(first_rate_values) != n
        or len(second_rate_values) != n
        or len(first_target_values) != n
        or len(second_target_values) != n
    ):
        return

    split_index = next((i for i, label in enumerate(date_labels) if "/" in label), len(date_labels))
    if split_index == 0:
        split_index = len(date_labels)

    summary_first = [first_rate_values[i] if i < split_index else None for i in range(n)]
    summary_second = [second_rate_values[i] if i < split_index else None for i in range(n)]
    daily_first = [first_rate_values[i] if i >= split_index else None for i in range(n)]
    daily_second = [second_rate_values[i] if i >= split_index else None for i in range(n)]

    helper_col = 70
    ws.cell(row=2, column=helper_col, value="序列")
    ws.cell(row=2, column=helper_col + 1, value="日期")
    ws.cell(row=3, column=helper_col, value="一次年/月汇总")
    ws.cell(row=4, column=helper_col, value="二次年/月汇总")
    ws.cell(row=5, column=helper_col, value="一次日趋势")
    ws.cell(row=6, column=helper_col, value="二次日趋势")
    ws.cell(row=7, column=helper_col, value="一次目標良率")
    ws.cell(row=8, column=helper_col, value="二次目標良率")

    for idx, row_vals in enumerate(
        zip(
            date_labels,
            summary_first,
            summary_second,
            daily_first,
            daily_second,
            first_target_values,
            second_target_values,
        )
    ):
        label, s1, s2, d1, d2, t1, t2 = row_vals
        col = helper_col + 1 + idx
        ws.cell(row=2, column=col, value=label)
        ws.cell(row=3, column=col, value=s1)
        ws.cell(row=4, column=col, value=s2)
        ws.cell(row=5, column=col, value=d1)
        ws.cell(row=6, column=col, value=d2)
        ws.cell(row=7, column=col, value=t1)
        ws.cell(row=8, column=col, value=t2)
        for r in range(3, 9):
            ws.cell(row=r, column=col).number_format = "0.00%"

    end_col = helper_col + len(date_labels)
    cats = Reference(ws, min_col=helper_col + 1, max_col=end_col, min_row=2, max_row=2)

    bar = BarChart()
    bar.title = "一次/二次良率趋势图"
    bar.type = "col"
    bar.style = 10
    bar.y_axis.title = "良率"
    bar.x_axis.title = "日期"
    bar.height = 5.0
    bar.width = 14.0
    bar.gapWidth = 140
    bar.y_axis.number_format = "0.00%"
    bar.title.txPr = _text_props(10)
    bar.x_axis.txPr = _text_props(8)
    bar.y_axis.txPr = _text_props(8)

    min_rate = min(first_rate_values + second_rate_values + first_target_values + second_target_values)
    bar.y_axis.scaling.min = max(0.90, min_rate - 0.01)
    bar.y_axis.scaling.max = 1.00

    bar_data = Reference(ws, min_col=helper_col, max_col=end_col, min_row=3, max_row=4)
    bar.add_data(bar_data, titles_from_data=True, from_rows=True)
    bar.set_categories(cats)
    bar.x_axis.delete = False
    bar.y_axis.delete = False
    bar.x_axis.tickLblPos = "nextTo"
    bar.y_axis.tickLblPos = "nextTo"
    if len(bar.series) >= 1:
        bar.series[0].graphicalProperties.solidFill = "8EC7F7"
        bar.series[0].graphicalProperties.line.solidFill = "4A90E2"
        bar.series[0].dLbls = DataLabelList()
        bar.series[0].dLbls.showVal = True
        bar.series[0].dLbls.showSerName = False
        bar.series[0].dLbls.showCatName = False
        bar.series[0].dLbls.showLegendKey = False
        bar.series[0].dLbls.showPercent = False
        bar.series[0].dLbls.dLblPos = "outEnd"
        bar.series[0].dLbls.numFmt = "0.00%"
        bar.series[0].dLbls.txPr = _text_props(7)
    if len(bar.series) >= 2:
        bar.series[1].graphicalProperties.solidFill = "A8E6CF"
        bar.series[1].graphicalProperties.line.solidFill = "27AE60"
        bar.series[1].dLbls = DataLabelList()
        bar.series[1].dLbls.showVal = True
        bar.series[1].dLbls.showSerName = False
        bar.series[1].dLbls.showCatName = False
        bar.series[1].dLbls.showLegendKey = False
        bar.series[1].dLbls.showPercent = False
        bar.series[1].dLbls.dLblPos = "outEnd"
        bar.series[1].dLbls.numFmt = "0.00%"
        bar.series[1].dLbls.txPr = _text_props(7)

    line = LineChart()
    line_data = Reference(ws, min_col=helper_col, max_col=end_col, min_row=5, max_row=8)
    line.add_data(line_data, titles_from_data=True, from_rows=True)
    line.set_categories(cats)
    line.y_axis.axId = 200
    line.y_axis.scaling.min = bar.y_axis.scaling.min
    line.y_axis.scaling.max = 1.00
    line.y_axis.number_format = "0.00%"
    line.y_axis.majorGridlines = None
    line.y_axis.delete = True

    last_idx = max(0, n - 1)
    if len(line.series) >= 1:
        line.series[0].graphicalProperties.line.solidFill = "F39C12"
        line.series[0].graphicalProperties.line.width = 28575
        line.series[0].marker.symbol = "circle"
        line.series[0].marker.size = 6
        line.series[0].dLbls = DataLabelList()
        line.series[0].dLbls.showVal = True
        line.series[0].dLbls.showSerName = False
        line.series[0].dLbls.showCatName = False
        line.series[0].dLbls.showLegendKey = False
        line.series[0].dLbls.showPercent = False
        line.series[0].dLbls.dLblPos = "b"
        line.series[0].dLbls.numFmt = "0.00%"
        line.series[0].dLbls.txPr = _text_props(7)
    if len(line.series) >= 2:
        line.series[1].graphicalProperties.line.solidFill = "1ABC9C"
        line.series[1].graphicalProperties.line.width = 28575
        line.series[1].marker.symbol = "circle"
        line.series[1].marker.size = 6
        line.series[1].dLbls = DataLabelList()
        line.series[1].dLbls.showVal = True
        line.series[1].dLbls.showSerName = False
        line.series[1].dLbls.showCatName = False
        line.series[1].dLbls.showLegendKey = False
        line.series[1].dLbls.showPercent = False
        line.series[1].dLbls.dLblPos = "t"
        line.series[1].dLbls.numFmt = "0.00%"
        line.series[1].dLbls.txPr = _text_props(7)
    if len(line.series) >= 3:
        line.series[2].graphicalProperties.line.solidFill = "D9534F"
        line.series[2].graphicalProperties.line.width = 19050
        line.series[2].graphicalProperties.line.dashStyle = "sysDot"
        line.series[2].marker.symbol = "triangle"
        line.series[2].marker.size = 5
        line.series[2].dLbls = DataLabelList()
        line.series[2].dLbls.showVal = False
        line.series[2].dLbls.showSerName = False
        line.series[2].dLbls.showCatName = False
        line.series[2].dLbls.showLegendKey = False
        line.series[2].dLbls.showPercent = False
        line.series[2].dLbls.dLblPos = "r"
        line.series[2].dLbls.numFmt = "0.00%"
        line.series[2].dLbls.txPr = _text_props(7)
        line.series[2].dLbls.dLbl = [
            DataLabel(
                idx=last_idx,
                showVal=True,
                showSerName=False,
                showCatName=False,
                showLegendKey=False,
                showPercent=False,
                showBubbleSize=False,
                dLblPos="r",
                numFmt="0.00%",
                separator="",
            )
        ]
    if len(line.series) >= 4:
        line.series[3].graphicalProperties.line.solidFill = "8E44AD"
        line.series[3].graphicalProperties.line.width = 19050
        line.series[3].graphicalProperties.line.dashStyle = "sysDot"
        line.series[3].marker.symbol = "triangle"
        line.series[3].marker.size = 5
        line.series[3].dLbls = DataLabelList()
        line.series[3].dLbls.showVal = False
        line.series[3].dLbls.showSerName = False
        line.series[3].dLbls.showCatName = False
        line.series[3].dLbls.showLegendKey = False
        line.series[3].dLbls.showPercent = False
        line.series[3].dLbls.dLblPos = "r"
        line.series[3].dLbls.numFmt = "0.00%"
        line.series[3].dLbls.txPr = _text_props(7)
        line.series[3].dLbls.dLbl = [
            DataLabel(
                idx=last_idx,
                showVal=True,
                showSerName=False,
                showCatName=False,
                showLegendKey=False,
                showPercent=False,
                showBubbleSize=False,
                dLblPos="r",
                numFmt="0.00%",
                separator="",
            )
        ]

    bar += line
    bar.legend.position = "b"
    bar.legend.txPr = _text_props(7.5)
    _finalize_dashboard_rate_chart(final_data, bar, ws)

def add_native_good_rate_chart_jinjia_cnc0(ws, final_data: pd.DataFrame, anchor: str = "A2") -> None:
    """
    openpyxl 组合图：双柱（一次/二次年或月汇总）+ 四线（两次日趋势 + 两项目标良率）。
    """
    def _text_props(size_pt: float) -> RichText:
        size = int(size_pt * 100)
        return RichText(
            bodyPr=RichTextProperties(),
            p=[
                Paragraph(
                    pPr=ParagraphProperties(defRPr=CharacterProperties(sz=size)),
                    endParaRPr=CharacterProperties(sz=size),
                )
            ],
        )

    date_labels = [str(col) for col in final_data.columns.tolist()[1:]]
    if not date_labels:
        return

    first_target_values = [_rate_to_float(v) for v in final_data.iloc[3][1:].tolist()]
    second_target_values = [_rate_to_float(v) for v in final_data.iloc[4][1:].tolist()]
    first_rate_values = [_rate_to_float(v) for v in final_data.iloc[8][1:].tolist()]
    second_rate_values = [_rate_to_float(v) for v in final_data.iloc[9][1:].tolist()]
    n = len(date_labels)
    if (
        not first_rate_values
        or len(first_rate_values) != n
        or len(second_rate_values) != n
        or len(first_target_values) != n
        or len(second_target_values) != n
    ):
        return

    split_index = next((i for i, label in enumerate(date_labels) if "/" in label), len(date_labels))
    if split_index == 0:
        split_index = len(date_labels)

    summary_first = [first_rate_values[i] if i < split_index else None for i in range(n)]
    summary_second = [second_rate_values[i] if i < split_index else None for i in range(n)]
    daily_first = [first_rate_values[i] if i >= split_index else None for i in range(n)]
    daily_second = [second_rate_values[i] if i >= split_index else None for i in range(n)]

    helper_col = 70
    ws.cell(row=2, column=helper_col, value="序列")
    ws.cell(row=2, column=helper_col + 1, value="日期")
    ws.cell(row=3, column=helper_col, value="一次年/月汇总")
    ws.cell(row=4, column=helper_col, value="二次年/月汇总")
    ws.cell(row=5, column=helper_col, value="一次日趋势")
    ws.cell(row=6, column=helper_col, value="二次日趋势")
    ws.cell(row=7, column=helper_col, value="一次目標良率")
    ws.cell(row=8, column=helper_col, value="二次目標良率")

    for idx, row_vals in enumerate(
        zip(
            date_labels,
            summary_first,
            summary_second,
            daily_first,
            daily_second,
            first_target_values,
            second_target_values,
        )
    ):
        label, s1, s2, d1, d2, t1, t2 = row_vals
        col = helper_col + 1 + idx
        ws.cell(row=2, column=col, value=label)
        ws.cell(row=3, column=col, value=s1)
        ws.cell(row=4, column=col, value=s2)
        ws.cell(row=5, column=col, value=d1)
        ws.cell(row=6, column=col, value=d2)
        ws.cell(row=7, column=col, value=t1)
        ws.cell(row=8, column=col, value=t2)
        for r in range(3, 9):
            ws.cell(row=r, column=col).number_format = "0.00%"

    end_col = helper_col + len(date_labels)
    cats = Reference(ws, min_col=helper_col + 1, max_col=end_col, min_row=2, max_row=2)

    bar = BarChart()
    bar.title = "一次/二次良率趋势图"
    bar.type = "col"
    bar.style = 10
    bar.y_axis.title = "良率"
    bar.x_axis.title = "日期"
    bar.height = 5.0
    bar.width = 14.0
    bar.gapWidth = 140
    bar.y_axis.number_format = "0.00%"
    bar.title.txPr = _text_props(10)
    bar.x_axis.txPr = _text_props(8)
    bar.y_axis.txPr = _text_props(8)

    min_rate = min(first_rate_values + second_rate_values + first_target_values + second_target_values)
    bar.y_axis.scaling.min = max(0.90, min_rate - 0.01)
    bar.y_axis.scaling.max = 1.00

    bar_data = Reference(ws, min_col=helper_col, max_col=end_col, min_row=3, max_row=4)
    bar.add_data(bar_data, titles_from_data=True, from_rows=True)
    bar.set_categories(cats)
    bar.x_axis.delete = False
    bar.y_axis.delete = False
    bar.x_axis.tickLblPos = "nextTo"
    bar.y_axis.tickLblPos = "nextTo"
    if len(bar.series) >= 1:
        bar.series[0].graphicalProperties.solidFill = "8EC7F7"
        bar.series[0].graphicalProperties.line.solidFill = "4A90E2"
        bar.series[0].dLbls = DataLabelList()
        bar.series[0].dLbls.showVal = True
        bar.series[0].dLbls.showSerName = False
        bar.series[0].dLbls.showCatName = False
        bar.series[0].dLbls.showLegendKey = False
        bar.series[0].dLbls.showPercent = False
        bar.series[0].dLbls.dLblPos = "outEnd"
        bar.series[0].dLbls.numFmt = "0.00%"
        bar.series[0].dLbls.txPr = _text_props(7)
    if len(bar.series) >= 2:
        bar.series[1].graphicalProperties.solidFill = "A8E6CF"
        bar.series[1].graphicalProperties.line.solidFill = "27AE60"
        bar.series[1].dLbls = DataLabelList()
        bar.series[1].dLbls.showVal = True
        bar.series[1].dLbls.showSerName = False
        bar.series[1].dLbls.showCatName = False
        bar.series[1].dLbls.showLegendKey = False
        bar.series[1].dLbls.showPercent = False
        bar.series[1].dLbls.dLblPos = "outEnd"
        bar.series[1].dLbls.numFmt = "0.00%"
        bar.series[1].dLbls.txPr = _text_props(7)

    line = LineChart()
    line_data = Reference(ws, min_col=helper_col, max_col=end_col, min_row=5, max_row=8)
    line.add_data(line_data, titles_from_data=True, from_rows=True)
    line.set_categories(cats)
    line.y_axis.axId = 200
    line.y_axis.scaling.min = bar.y_axis.scaling.min
    line.y_axis.scaling.max = 1.00
    line.y_axis.number_format = "0.00%"
    line.y_axis.majorGridlines = None
    line.y_axis.delete = True

    last_idx = max(0, n - 1)
    if len(line.series) >= 1:
        line.series[0].graphicalProperties.line.solidFill = "F39C12"
        line.series[0].graphicalProperties.line.width = 28575
        line.series[0].marker.symbol = "circle"
        line.series[0].marker.size = 6
        line.series[0].dLbls = DataLabelList()
        line.series[0].dLbls.showVal = True
        line.series[0].dLbls.showSerName = False
        line.series[0].dLbls.showCatName = False
        line.series[0].dLbls.showLegendKey = False
        line.series[0].dLbls.showPercent = False
        line.series[0].dLbls.dLblPos = "b"
        line.series[0].dLbls.numFmt = "0.00%"
        line.series[0].dLbls.txPr = _text_props(7)
    if len(line.series) >= 2:
        line.series[1].graphicalProperties.line.solidFill = "1ABC9C"
        line.series[1].graphicalProperties.line.width = 28575
        line.series[1].marker.symbol = "circle"
        line.series[1].marker.size = 6
        line.series[1].dLbls = DataLabelList()
        line.series[1].dLbls.showVal = True
        line.series[1].dLbls.showSerName = False
        line.series[1].dLbls.showCatName = False
        line.series[1].dLbls.showLegendKey = False
        line.series[1].dLbls.showPercent = False
        line.series[1].dLbls.dLblPos = "t"
        line.series[1].dLbls.numFmt = "0.00%"
        line.series[1].dLbls.txPr = _text_props(7)
    if len(line.series) >= 3:
        line.series[2].graphicalProperties.line.solidFill = "D9534F"
        line.series[2].graphicalProperties.line.width = 19050
        line.series[2].graphicalProperties.line.dashStyle = "sysDot"
        line.series[2].marker.symbol = "triangle"
        line.series[2].marker.size = 5
        line.series[2].dLbls = DataLabelList()
        line.series[2].dLbls.showVal = False
        line.series[2].dLbls.showSerName = False
        line.series[2].dLbls.showCatName = False
        line.series[2].dLbls.showLegendKey = False
        line.series[2].dLbls.showPercent = False
        line.series[2].dLbls.dLblPos = "r"
        line.series[2].dLbls.numFmt = "0.00%"
        line.series[2].dLbls.txPr = _text_props(7)
        line.series[2].dLbls.dLbl = [
            DataLabel(
                idx=last_idx,
                showVal=True,
                showSerName=False,
                showCatName=False,
                showLegendKey=False,
                showPercent=False,
                showBubbleSize=False,
                dLblPos="r",
                numFmt="0.00%",
                separator="",
            )
        ]
    if len(line.series) >= 4:
        line.series[3].graphicalProperties.line.solidFill = "8E44AD"
        line.series[3].graphicalProperties.line.width = 19050
        line.series[3].graphicalProperties.line.dashStyle = "sysDot"
        line.series[3].marker.symbol = "triangle"
        line.series[3].marker.size = 5
        line.series[3].dLbls = DataLabelList()
        line.series[3].dLbls.showVal = False
        line.series[3].dLbls.showSerName = False
        line.series[3].dLbls.showCatName = False
        line.series[3].dLbls.showLegendKey = False
        line.series[3].dLbls.showPercent = False
        line.series[3].dLbls.dLblPos = "r"
        line.series[3].dLbls.numFmt = "0.00%"
        line.series[3].dLbls.txPr = _text_props(7)
        line.series[3].dLbls.dLbl = [
            DataLabel(
                idx=last_idx,
                showVal=True,
                showSerName=False,
                showCatName=False,
                showLegendKey=False,
                showPercent=False,
                showBubbleSize=False,
                dLblPos="r",
                numFmt="0.00%",
                separator="",
            )
        ]

    bar += line
    bar.legend.position = "b"
    bar.legend.txPr = _text_props(7.5)
    _finalize_dashboard_rate_chart(final_data, bar, ws)

def build_dashboard_export_bytes(
    key,
    process,
    data: pd.DataFrame,
    production_name: str | None = None,
    inspection_location: str | None = None,
    shift: str | None = None,
) -> bytes:
    sel_p = production_name if production_name in PRODUCT_NAME_OPTIONS else DEFAULT_PRODUCT_NAME
    sel_i = (
        inspection_location
        if inspection_location in INSPECTION_LOCATION_OPTIONS
        else "不限"
    )
    sel_shift = shift if shift in SHIFT_OPTIONS else "不限"
    title_parts = [str(key or "-"), str(process or "-"), f"品名 {sel_p}", f"班别 {sel_shift}"]
    if (process or "") == "CNC0" and sel_i != "不限":
        title_parts.append(f"抽检 {sel_i}")
    title = " | ".join(title_parts)

    # 初始化变量
    final_data = None

    if key == "沖壓" and process == "鍛壓" and not data.empty:
        d = data.assign(_day=pd.to_datetime(data["production_date"]).dt.normalize())
        num_cols = ["input_total", "bad_total", *[f"{a}_badnum_total" for a, _ in DUANYA_DEFECT_TYPES]]
        for c in num_cols:
            d[c] = pd.to_numeric(d[c], errors="coerce").fillna(0)
        d = d[["_day", *num_cols]]
        # 按天合并（仅日期 + 数值列）
        d = d.groupby("_day", as_index=False, sort=True).sum(numeric_only=True)
        cn = {"_day": "生產日期", "input_total": "總投入數", "bad_total": "總不良數"}
        cn |= {f"{a}_badnum_total": z for a, z in DUANYA_DEFECT_TYPES}
        d = d.rename(columns=cn)
        d["生產日期"] = pd.to_datetime(d["生產日期"]) 

        data = d[["生產日期", "總投入數", "總不良數", *[z for _, z in DUANYA_DEFECT_TYPES]]]

        defect_cols = [z for _, z in DUANYA_DEFECT_TYPES]
        defect_list = data[defect_cols].sum().sort_values(ascending=False).index.tolist()

        data = data[["生產日期", "總投入數", "總不良數"] + defect_list]
        val_cols = ["總投入數", "總不良數"] + defect_list

        # 月小计
        monthly = data.groupby(data["生產日期"].dt.to_period("M"))[val_cols].sum().reset_index()
        monthly["生產日期"] = monthly["生產日期"].dt.month.astype(str) + "月"  

        # 年小计
        yearly = data.groupby(data["生產日期"].dt.year)[val_cols].sum().reset_index()
        yearly["生產日期"] = yearly["生產日期"].astype(str) + "年"

        data["生產日期"] = pd.to_datetime(d["生產日期"]) .dt.strftime("%m/%d")
        # 拼接
        final_data = pd.concat([yearly, monthly,data], ignore_index=True)[["生產日期"] + val_cols]
        final_data["良品數"] = final_data["總投入數"]-final_data["總不良數"]
        final_data["目標良率"] = "99.80%"
        final_data["總良率"] = (final_data["良品數"] / final_data["總投入數"]).map(lambda x: f"{x:.2%}")
        final_data = final_data[["生產日期","總投入數","良品數","目標良率","總不良數","總良率"]+defect_list]
        # 在转置后，将索引（原列名）作为第一列
        final_data = final_data.T
        new_header = final_data.iloc[0]  # 第一行作为新列名
        final_data = final_data[1:].copy()  # 去掉第一行并复制
        final_data.columns = new_header.values  # 使用原来的第一行值作为新列名
        dates = ["總投入數","良品數","目標良率","總不良數","總良率"] + defect_list
        final_data.insert(0, '項目\日期', dates)  # 在位置0插入'日期'列
        final_data.index = range(len(final_data))  # 重置索引为数字

    if key == "沖壓" and process == "固熔" and not data.empty:  

        d = data.assign(_day=pd.to_datetime(data["production_date"]).dt.normalize())
        num_cols = ["input_total", "bad_total", *[f"{a}_badnum_total" for a, _ in GURONG_DEFECT_TYPES]]
        for c in num_cols:
            d[c] = pd.to_numeric(d[c], errors="coerce").fillna(0)
        d = d[["_day", *num_cols]]
        # 按天合并（仅日期 + 数值列）
        d = d.groupby("_day", as_index=False, sort=True).sum(numeric_only=True)
        cn = {"_day": "生產日期", "input_total": "總投入數", "bad_total": "總不良數"}
        cn |= {f"{a}_badnum_total": z for a, z in GURONG_DEFECT_TYPES}
        d = d.rename(columns=cn)
        d["生產日期"] = pd.to_datetime(d["生產日期"]) 

        data = d[["生產日期", "總投入數", "總不良數", *[z for _, z in GURONG_DEFECT_TYPES]]]

        defect_cols = [z for _, z in GURONG_DEFECT_TYPES]
        defect_list = data[defect_cols].sum().sort_values(ascending=False).index.tolist()

        data = data[["生產日期", "總投入數", "總不良數"] + defect_list]
        val_cols = ["總投入數", "總不良數"] + defect_list

        # 月小计
        monthly = data.groupby(data["生產日期"].dt.to_period("M"))[val_cols].sum().reset_index()
        monthly["生產日期"] = monthly["生產日期"].dt.month.astype(str) + "月"  

        # 年小计
        yearly = data.groupby(data["生產日期"].dt.year)[val_cols].sum().reset_index()
        yearly["生產日期"] = yearly["生產日期"].astype(str) + "年"

        data["生產日期"] = pd.to_datetime(d["生產日期"]) .dt.strftime("%m/%d")
        # 拼接
        final_data = pd.concat([yearly, monthly,data], ignore_index=True)[["生產日期"] + val_cols]
        final_data["良品數"] = final_data["總投入數"]-final_data["總不良數"]
        final_data["目標良率"] = "100.00%"
        final_data["總良率"] = (final_data["良品數"] / final_data["總投入數"]).map(lambda x: f"{x:.2%}")
        final_data = final_data[["生產日期","總投入數","良品數","目標良率","總不良數","總良率"]+defect_list]
        # 在转置后，将索引（原列名）作为第一列
        final_data = final_data.T
        new_header = final_data.iloc[0]  # 第一行作为新列名
        final_data = final_data[1:].copy()  # 去掉第一行并复制
        final_data.columns = new_header.values  # 使用原来的第一行值作为新列名
        dates = ["總投入數","良品數","目標良率","總不良數","總良率"] + defect_list
        final_data.insert(0, '項目\日期', dates)  # 在位置0插入'日期'列
        final_data.index = range(len(final_data))  # 重置索引为数字

    if key == "沖壓" and process == "時效" and not data.empty:  

        d = data.assign(_day=pd.to_datetime(data["production_date"]).dt.normalize())
        num_cols = ["input_total", "bad_total", *[f"{a}_badnum_total" for a, _ in SHIXIAO_DEFECT_TYPES]]
        for c in num_cols:
            d[c] = pd.to_numeric(d[c], errors="coerce").fillna(0)
        d = d[["_day", *num_cols]]
        # 按天合并（仅日期 + 数值列）
        d = d.groupby("_day", as_index=False, sort=True).sum(numeric_only=True)
        cn = {"_day": "生產日期", "input_total": "總投入數", "bad_total": "總不良數"}
        cn |= {f"{a}_badnum_total": z for a, z in SHIXIAO_DEFECT_TYPES}
        d = d.rename(columns=cn)
        d["生產日期"] = pd.to_datetime(d["生產日期"]) 

        data = d[["生產日期", "總投入數", "總不良數", *[z for _, z in SHIXIAO_DEFECT_TYPES]]]

        defect_cols = [z for _, z in SHIXIAO_DEFECT_TYPES]
        defect_list = data[defect_cols].sum().sort_values(ascending=False).index.tolist()

        data = data[["生產日期", "總投入數", "總不良數"] + defect_list]
        val_cols = ["總投入數", "總不良數"] + defect_list

        # 月小计
        monthly = data.groupby(data["生產日期"].dt.to_period("M"))[val_cols].sum().reset_index()
        monthly["生產日期"] = monthly["生產日期"].dt.month.astype(str) + "月"  

        # 年小计
        yearly = data.groupby(data["生產日期"].dt.year)[val_cols].sum().reset_index()
        yearly["生產日期"] = yearly["生產日期"].astype(str) + "年"

        data["生產日期"] = pd.to_datetime(d["生產日期"]) .dt.strftime("%m/%d")
        # 拼接
        final_data = pd.concat([yearly, monthly,data], ignore_index=True)[["生產日期"] + val_cols]
        final_data["良品數"] = final_data["總投入數"]-final_data["總不良數"]
        final_data["目標良率"] = "100.00%"
        final_data["總良率"] = (final_data["良品數"] / final_data["總投入數"]).map(lambda x: f"{x:.2%}")
        final_data = final_data[["生產日期","總投入數","良品數","目標良率","總不良數","總良率"]+defect_list]
        # 在转置后，将索引（原列名）作为第一列
        final_data = final_data.T
        new_header = final_data.iloc[0]  # 第一行作为新列名
        final_data = final_data[1:].copy()  # 去掉第一行并复制
        final_data.columns = new_header.values  # 使用原来的第一行值作为新列名
        dates = ["總投入數","良品數","目標良率","總不良數","總良率"] + defect_list
        final_data.insert(0, '項目\日期', dates)  # 在位置0插入'日期'列
        final_data.index = range(len(final_data))  # 重置索引为数字

    if key == "金加" and process == "CNC0 全檢" and not data.empty: 

        d = data.assign(_day=pd.to_datetime(data["production_date"]).dt.normalize())
        num_cols = ["input_total", "bad_total","reworkable_bad","unreworkable_bad",
        *[f"{a}_badnum_total" for a, _ in CNC0_FULL_DEFECT_TYPES]]
        for c in num_cols:
            d[c] = pd.to_numeric(d[c], errors="coerce").fillna(0)
        d = d[["_day", *num_cols]]
        # 按天合并（仅日期 + 数值列）
        d = d.groupby("_day", as_index=False, sort=True).sum(numeric_only=True)
        cn = {"_day": "生產日期", "input_total": "總投入數", "bad_total": "總不良數",
        "reworkable_bad":"可重工不良數","unreworkable_bad":"不可重工不良數"}

        cn |= {f"{a}_badnum_total": z for a, z in CNC0_FULL_DEFECT_TYPES}
        d = d.rename(columns=cn)
        d["生產日期"] = pd.to_datetime(d["生產日期"]) 

        data = d[["生產日期", "總投入數", "總不良數","可重工不良數", "不可重工不良數",
        *[z for _, z in CNC0_FULL_DEFECT_TYPES]]]

        defect_cols = [z for _, z in CNC0_FULL_DEFECT_TYPES]
        defect_list = data[defect_cols].sum().sort_values(ascending=False).index.tolist()

        data = data[["生產日期", "總投入數", "總不良數","可重工不良數", "不可重工不良數"] + defect_list]
        val_cols = ["總投入數", "總不良數","可重工不良數", "不可重工不良數"] + defect_list

        # 月小计
        monthly = data.groupby(data["生產日期"].dt.to_period("M"))[val_cols].sum().reset_index()
        monthly["生產日期"] = monthly["生產日期"].dt.month.astype(str) + "月"  

        # 年小计
        yearly = data.groupby(data["生產日期"].dt.year)[val_cols].sum().reset_index()
        yearly["生產日期"] = yearly["生產日期"].astype(str) + "年"

        data["生產日期"] = pd.to_datetime(d["生產日期"]) .dt.strftime("%m/%d")
        # 拼接
        final_data = pd.concat([yearly, monthly,data], ignore_index=True)[["生產日期"] + val_cols]
        final_data["良品數"] = final_data["總投入數"]-final_data["總不良數"]
        final_data["一次目標良率"] = "99.90%"
        final_data["二次目標良率"] = "100.00%"
        final_data["一次良率"] = (final_data["良品數"] / final_data["總投入數"]).map(lambda x: f"{x:.2%}")
        final_data["二次良率"] = ((final_data["總投入數"]-final_data["不可重工不良數"]) / final_data["總投入數"]).map(lambda x: f"{x:.2%}")
        final_data = final_data[["生產日期","總投入數","良品數","一次目標良率","二次目標良率",
        "總不良數","可重工不良數", "不可重工不良數","一次良率","二次良率"]+defect_list]
        # 在转置后，将索引（原列名）作为第一列
        final_data = final_data.T
        new_header = final_data.iloc[0]  # 第一行作为新列名
        final_data = final_data[1:].copy()  # 去掉第一行并复制
        final_data.columns = new_header.values  # 使用原来的第一行值作为新列名
        dates = ["總投入數","良品數","一次目標良率","二次目標良率",
        "總不良數","可重工不良數", "不可重工不良數","一次良率","二次良率"] + defect_list
        final_data.insert(0, '項目\日期', dates)  # 在位置0插入'日期'列
        final_data.index = range(len(final_data))  # 重置索引为数字

    if key == "金加" and process == "CNC0" and not data.empty:  

        # 这里因为数据库没有合并一次、二次特定不良类型的数据，所以先合并！
        for prefix, _ in CNC0_DEFECT_TYPES:
            rw_col = f"{prefix}_badnum_reworkable"
            urw_col = f"{prefix}_badnum_unreworkable"
            total_col = f"{prefix}_badnum_total"
            rw = pd.to_numeric(data.get(rw_col, 0), errors="coerce").fillna(0)
            urw = pd.to_numeric(data.get(urw_col, 0), errors="coerce").fillna(0)
            data[total_col] = rw + urw


        d = data.assign(_day=pd.to_datetime(data["production_date"]).dt.normalize())
        num_cols = ["input_total","sample","bad_total","reworkable_bad","unreworkable_bad",
        *[f"{a}_badnum_total" for a, _ in CNC0_DEFECT_TYPES]]
        for c in num_cols:
            d[c] = pd.to_numeric(d[c], errors="coerce").fillna(0)
        d = d[["_day", *num_cols]]
        # 按天合并（仅日期 + 数值列）
        d = d.groupby("_day", as_index=False, sort=True).sum(numeric_only=True)
        cn = {"_day": "生產日期", "input_total": "總投入數", "sample":"抽检数","bad_total": "總不良數",
        "reworkable_bad":"可重工不良數","unreworkable_bad":"不可重工不良數"}

        cn |= {f"{a}_badnum_total": z for a, z in CNC0_DEFECT_TYPES}
        d = d.rename(columns=cn)
        d["生產日期"] = pd.to_datetime(d["生產日期"]) 

        data = d[["生產日期", "總投入數", "抽检数", "總不良數","可重工不良數", "不可重工不良數",
        *[z for _, z in CNC0_DEFECT_TYPES]]]

        defect_cols = [z for _, z in CNC0_DEFECT_TYPES]
        defect_list = data[defect_cols].sum().sort_values(ascending=False).index.tolist()

        data = data[["生產日期", "總投入數", "抽检数","總不良數","可重工不良數", "不可重工不良數"] + defect_list]
        val_cols = ["總投入數","抽检数", "總不良數","可重工不良數", "不可重工不良數"] + defect_list

        # 月小计
        monthly = data.groupby(data["生產日期"].dt.to_period("M"))[val_cols].sum().reset_index()
        monthly["生產日期"] = monthly["生產日期"].dt.month.astype(str) + "月"  

        # 年小计
        yearly = data.groupby(data["生產日期"].dt.year)[val_cols].sum().reset_index()
        yearly["生產日期"] = yearly["生產日期"].astype(str) + "年"

        data["生產日期"] = pd.to_datetime(d["生產日期"]) .dt.strftime("%m/%d")
        # 拼接
        final_data = pd.concat([yearly, monthly,data], ignore_index=True)[["生產日期"] + val_cols]
        final_data["良品數"] = final_data["總投入數"]-final_data["總不良數"]
        final_data["一次目標良率"] = "99.70%"
        final_data["二次目標良率"] = "100.00%"
        final_data["一次良率"] = (final_data["良品數"] / final_data["總投入數"]).map(lambda x: f"{x:.2%}")
        final_data["二次良率"] = ((final_data["總投入數"]-final_data["不可重工不良數"]) / final_data["總投入數"]).map(lambda x: f"{x:.2%}")
        final_data = final_data[["生產日期","總投入數","抽检数","良品數","一次目標良率","二次目標良率",
        "總不良數","可重工不良數", "不可重工不良數","一次良率","二次良率"]+defect_list]
        # 在转置后，将索引（原列名）作为第一列
        final_data = final_data.T
        new_header = final_data.iloc[0]  # 第一行作为新列名
        final_data = final_data[1:].copy()  # 去掉第一行并复制
        final_data.columns = new_header.values  # 使用原来的第一行值作为新列名
        dates = ["總投入數","抽检数","良品數","一次目標良率","二次目標良率",
        "總不良數","可重工不良數", "不可重工不良數","一次良率","二次良率"] + defect_list
        final_data.insert(0, '項目\日期', dates)  # 在位置0插入'日期'列
        final_data.index = range(len(final_data))  # 重置索引为数字

        
    # 往excel填数据
    buf = BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as w:
        if final_data is not None:
            final_data.to_excel(w, index=False, sheet_name="数据看板", startrow=10,startcol=1)
            ws = w.sheets["数据看板"]
        else:
            ws = w.book.create_sheet("数据看板")
        ws["A1"] = title

        if final_data is not None:
            # 拉高图表区域（第2~10行），避免图像被压扁导致内容难以辨识
            for r in range(2, 11):
                current_height = ws.row_dimensions[r].height or 15
                ws.row_dimensions[r].height = max(current_height, 28)
            if key == "沖壓":
                add_native_good_rate_chart_choya(ws, final_data, anchor="A2")
            elif key == "金加" and process == "CNC0 全檢":
                add_native_good_rate_chart_jinjia_cnc0_full(ws, final_data, anchor="A2")
            else: #cnc0
                add_native_good_rate_chart_jinjia_cnc0(ws, final_data, anchor="A2")


        process_x_map = {
            "鍛壓": 17,
            "固熔": 17,   
            "時效": 17,
            "CNC0 全檢": 21, 
            "CNC0": 22,     
        }

        x = process_x_map.get(process, 17)
        ws[f"A{x}"] = "前五项不良"
            
       
        
        # 计算数据范围（从startrow+1开始，因为标题占一行）
        start_row = 11  # startrow + 1（跳过标题行）
        start_col = 1   # startcol + 1（A列是1，B列是2，以此类推）
        end_row = start_row + len(final_data) if final_data is not None else start_row
        end_col = start_col + len(final_data.columns) if final_data is not None else start_col

        # 让特定单元格的文字竖放（A17单元格）
        cell = ws[f"A{x}"]
        cell.alignment = Alignment(text_rotation=255, vertical='center', horizontal='center')
        
        # 合并A和B列的对应行，但跳过A{x}和B{x}行
        for row in range(start_row, x):
            if row != x:  # 跳过第x行
                # 先将B列的内容复制到A列（仅当final_data存在时）
                if final_data is not None and row < start_row + len(final_data) + 1:
                    ws.cell(row=row, column=1).value = ws.cell(row=row, column=2).value
                # 然后再合并A和B列
                ws.merge_cells(f'A{row}:B{row}')

        ws.merge_cells(f"A{x}:A{x+4}")

         
        # 定义细边框
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # 为数据区域添加边框
        for row in range(start_row, end_row + 1):
            for col in range(start_col, end_col + 1):
                ws.cell(row=row, column=col).border = thin_border

        # 自动调整列宽以适应内容
        for col_idx in range(start_col, end_col + 1):
            max_length = 0
            column_letter = ws.cell(row=1, column=col_idx).column_letter  # 获取列字母
            
            for row_idx in range(start_row, end_row + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                try:
                    if cell.value is not None:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)  # 最大宽度限制为50，避免过宽
            ws.column_dimensions[column_letter].width = adjusted_width
                
    return buf.getvalue()
